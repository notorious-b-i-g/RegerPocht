import tkinter as tk
from tkinter import simpledialog
from selenium import webdriver
from selenium.webdriver.common.by import By
import random
import string
import time
import openpyxl
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys

data = []
#Создайте экземпляр драйвера
driver = webdriver.Chrome()

#Перейти на нужную страницу
driver.get('https://passport.yandex.ru/auth/reg/portal?retpath=https%3A%2F%2Fmail.yandex.ru')  # Замените на URL вашей страницы

#Добавляем небольшую задержку для уверенности в загрузке элементов
time.sleep(2)  # Задержка в 2 секунды

#Найти элемент по классу и ввести данные
def phone_input():
    global phone_number
    root = tk.Tk()
    root.withdraw()
    phone_number = simpledialog.askstring("Ввод", "Введите ваш номер телефона:", parent=root)
    phone_input_field = driver.find_element(By.CLASS_NAME, 'Textinput-Control')
    phone_input_field.send_keys(phone_number)
    button = driver.find_element(By.CLASS_NAME, 'Button2_size_xxl')
    button.click()
def code_input():
    root = tk.Tk()
    root.withdraw()
    verification_code = simpledialog.askstring("Ввод", "Введите код подтверждения:", parent=root)
    code_enter = driver.find_element(By.ID, 'passp-field-phoneCode')
    code_enter.send_keys(verification_code)
def read_names_from_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return [line.strip() for line in file.readlines()]



def generate_russian_name():
    #"""Генерирует случайное русское имя и фамилию."""
    first_names = read_names_from_file('first_names.txt')
    last_names = read_names_from_file('last_names.txt')
    first_name = random.choice(first_names)
    last_name = random.choice(last_names)
    return first_name, last_name

def names_input(a, b):
    time.sleep(2)
    first_name_input = driver.find_element(By.ID, 'passp-field-firstname')
    second_name_input = driver.find_element(By.ID, 'passp-field-lastname')
    print(first_name_input)
    print(second_name_input)



    first_name_input.send_keys(a)
    second_name_input.send_keys(b)
    button = driver.find_element(By.CLASS_NAME, 'Button2_size_xxl')
    button.click()
    time.sleep(2)


def transliterate_to_latin(text):
    #"""Транслитерация русского текста в латиницу."""
    translation_table = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
        'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
    }
    return ''.join(translation_table.get(char, char) for char in text)

def generate_random_email_login(first_name, last_name, min_length=16, max_length=30):
    global email_login
    #"""Генерирует случайный логин для почты с минимальной длиной 16 символов."""
    # Транслитерация и преобразование в нижний регистр
    first_name = first_name.lower()
    last_name = last_name.lower()

    while True:
        # Случайные числа и буквы
        numbers = str(random.randint(100, 999))
        random_letters = ''.join(random.choices(string.ascii_lowercase, k=5))

        # Случайная комбинация элементов
        elements = [first_name, last_name, random_letters]
        random.shuffle(elements)
        email_login = ''.join(numbers).join(elements)

        # Проверяем, достигнута ли минимальная длина
        if len(email_login) >= min_length:
            break
    if len(email_login) > max_length:
        email_login = email_login[0:29]
    login_input = driver.find_element(By.ID, 'passp-field-login')
    login_input.send_keys(Keys.CONTROL + "a")
    login_input.send_keys(Keys.DELETE)
    login_input.send_keys(email_login)
    time.sleep(2)
    button = driver.find_element(By.ID, 'passp:Login:registration')
    button.click()


def password_create():
    global password
    #"""Генерирует случайный пароль заданной длины."""
    length = 9
    password_input = driver.find_element(By.ID, 'passp-field-password')
    characters = string.ascii_letters + string.digits + string.punctuation
    password = ''.join(random.choice(characters) for i in range(length))
    password_input.send_keys(password)
    print(password)
    button = driver.find_element(By.CLASS_NAME, 'Button2_size_xxl')
    button.click()
    time.sleep(1)
    button = driver.find_element(By.CLASS_NAME, 'Button2_size_xxl')
    button.click()

def data_write_in():
    global first, last, email_login, password, phone_number
    file_path = 'pochta.xlsx'

    # Проверяем, существует ли файл
    try:
        wb = load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        # Если файл не существует, создаем его
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Имя", "Фамилия", "Логин (почта)", "Пароль", "Телефон"])  # Заголовки

    # Добавляем новую строку с данными
    ws.append([first, last, email_login, password, phone_number])

    # Сохраняем файл
    wb.save(file_path)
def new_profile():
    driver.get('https://passport.yandex.ru/auth/reg/portal?retpath=https%3A%2F%2Fmail.yandex.ru')
    time.sleep(6)

while True:
    phone_input()
    code_input()
    first, last = generate_russian_name()

    names_input(first, last)
    generate_random_email_login(transliterate_to_latin(first), transliterate_to_latin(last))
    password_create()
    data_write_in()
    new_profile()
